"""
data_loader.py
Reads calculated values from the 180DC ESCP Excel financial model.
All row/column references are anchored to v3 of the model.
"""

import os
from pathlib import Path
import streamlit as st
from openpyxl import load_workbook

# ── File path ────────────────────────────────────────────────────────────────
DATA_DIR  = Path(__file__).parent / "data"
XLSX_FILE = DATA_DIR / "180DC_ESCP_FinancialModel_Fall2026_v4.xlsx"

# ── Cell maps ────────────────────────────────────────────────────────────────
# Assumption sheet — (row, col=2)
ASSUMPTION_CELLS = {
    "total_members":          5,
    "active_members":         6,
    "exec_members":           7,
    "campuses":               8,
    "consultants_per_proj":   9,
    "semester_months":        10,
    "projects":               12,
    "symbolic_fee":           13,
    "contribution_rate":      14,
    "project_revenue":        15,
    "num_sponsors":           17,
    "avg_sponsor_value":      18,
    "sponsorship_revenue":    19,
    "escp_grant":             21,
    "membership_fee":         22,
    "gross_membership":       23,
    "event_ticket_rev":       24,
    "other_revenue":          25,
    "num_events":             27,
    "cost_per_event":         28,
    "total_event_cost":       29,
    "marketing_pm":           31,
    "software_pm":            32,
    "admin_pm":               33,
    "delivery_cost_per_proj": 35,
    "travel_cost":            36,
    "contingency_pct":        38,
    "target_margin":          39,
}

# P&L Model sheet — col 3=Sep, 4=Oct, 5=Nov, 6=Dec, 7=Total
PL_ROWS = {
    "project_fees":   6,
    "sponsorship":    7,
    "escp_grant":     8,
    "membership":     9,
    "event_revenue":  10,
    "other_revenue":  11,
    "total_revenue":  12,
    "events_cost":    15,
    "proj_delivery":  16,
    "travel":         17,
    "marketing":      18,
    "software":       19,
    "admin":          20,
    "contingency":    21,
    "total_costs":    22,
    "net_surplus":    25,
    "cumulative":     26,
    "net_margin":     27,
    "cost_ratio":     28,
}

MONTHS = ["September", "October", "November", "December"]
MONTH_COLS = [3, 4, 5, 6]
TOTAL_COL  = 7

# KPI sheet — col 3=name, 4=value, 6=target, 7=status
KPI_ROWS = list(range(5, 13))   # rows 5–12
KPI_NAMES = [
    "Sponsorship Coverage Ratio (%)",
    "Cost per Member (€)",
    "Net Surplus / (Deficit) (€)",
    "Net Margin (%)",
    "Revenue per Project (€)",
    "Institutional & Sponsorship as % of Revenue",
    "Events Cost as % of Total Costs",
    "Revenue per Active Member (€)",
]
KPI_TARGETS = [
    "≥ 60%", "≤ €100", "> €0", "≥ 5%",
    "≥ €1,200", "≥ 50%", "≤ 35%", "≥ €200",
]
KPI_FORMATS = [
    "pct", "eur", "eur", "pct",
    "eur", "pct", "pct", "eur",
]


# ── Loader ────────────────────────────────────────────────────────────────────
@st.cache_data(ttl=60)
def load_model(path: str = str(XLSX_FILE)) -> dict:
    """
    Load all data from the Excel model.
    Returns a nested dict: {assumptions, pl_monthly, pl_total, kpis}.
    """
    if not Path(path).exists():
        raise FileNotFoundError(
            f"Excel model not found at: {path}\n"
            "Place '180DC_ESCP_FinancialModel_Fall2026_v4.xlsx' inside the /data folder."
        )

    wb = load_workbook(path, data_only=True)

    # ── Assumptions ──────────────────────────────────────────────────────────
    ws_a = wb["Assumptions"]
    assumptions = {
        key: (ws_a.cell(row=row, column=2).value or 0)
        for key, row in ASSUMPTION_CELLS.items()
    }

    # ── P&L monthly + totals ─────────────────────────────────────────────────
    ws_m = wb["P&L Model"]

    def read_row(row_num: int) -> dict:
        monthly = [
            ws_m.cell(row=row_num, column=col).value or 0
            for col in MONTH_COLS
        ]
        total = ws_m.cell(row=row_num, column=TOTAL_COL).value or 0
        return {"monthly": monthly, "total": total}

    pl = {key: read_row(row) for key, row in PL_ROWS.items()}

    # Derived monthly net surplus (in case cumulative is needed)
    pl["monthly_surplus"] = {
        "monthly": [
            (pl["total_revenue"]["monthly"][i] - pl["total_costs"]["monthly"][i])
            for i in range(4)
        ],
        "total": pl["net_surplus"]["total"],
    }

    pl["cumulative_surplus"] = {
        "monthly": _running_total(pl["monthly_surplus"]["monthly"]),
        "total":   pl["net_surplus"]["total"],
    }

    # ── KPIs ─────────────────────────────────────────────────────────────────
    ws_k = wb["KPIs"]
    kpis = []
    for i, row in enumerate(KPI_ROWS):
        value  = ws_k.cell(row=row, column=4).value or 0
        status = ws_k.cell(row=row, column=7).value or "-"
        kpis.append({
            "name":   KPI_NAMES[i],
            "value":  value,
            "target": KPI_TARGETS[i],
            "status": status,
            "format": KPI_FORMATS[i],
        })

    wb.close()
    return {"assumptions": assumptions, "pl": pl, "kpis": kpis}


def _running_total(values: list) -> list:
    out, acc = [], 0
    for v in values:
        acc += v
        out.append(acc)
    return out


# ── Project Tracker loader ────────────────────────────────────────────────────
PROJECT_STATUS_ROWS = range(5, 13)   # rows 5–12, 8 projects
PROJECT_COLS = {
    "num": 2, "client": 3, "sector": 4, "leader": 5,
    "team_size": 7, "start": 8, "end": 9, "status": 10,
    "hrs_budget": 11, "hrs_actual": 12, "completion_pct": 13,
    "fee": 14, "invoiced": 15, "received": 16, "outstanding": 17,
    "notes": 18,
}

# Scenario tab cell references (col 3=Conservative, 4=Base, 5=Optimistic)
SCENARIO_OUTCOME_ROWS = {
    "total_revenue": 20,
    "total_costs":   21,
    "net_surplus":   22,
    "net_margin":    23,
}

# Variance tab — Spring vs Fall
VARIANCE_ROWS = {
    "project_fees":  6,
    "sponsorship":   7,
    "escp_grant":    8,
    "membership":    9,
    "event_revenue": 10,
    "other_revenue": 11,
    "total_revenue": 12,
    "events_cost":   14,
    "proj_delivery": 15,
    "travel":        16,
    "marketing":     17,
    "software":      18,
    "admin":         19,
    "contingency":   20,
    "total_costs":   21,
    "net_surplus":   22,
    "net_margin":    23,
}


@st.cache_data(ttl=60)
def load_projects(path: str = str(XLSX_FILE)) -> list:
    wb = load_workbook(path, data_only=True)
    ws = wb["Project Tracker"]
    projects = []
    for r in PROJECT_STATUS_ROWS:
        projects.append({
            k: ws.cell(row=r, column=c).value
            for k, c in PROJECT_COLS.items()
        })
    wb.close()
    return projects


@st.cache_data(ttl=60)
def load_scenarios(path: str = str(XLSX_FILE)) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb["Scenarios"]
    out = {}
    for key, row in SCENARIO_OUTCOME_ROWS.items():
        out[key] = {
            "conservative": ws.cell(row=row, column=3).value or 0,
            "base":         ws.cell(row=row, column=4).value or 0,
            "optimistic":   ws.cell(row=row, column=5).value or 0,
        }
    wb.close()
    return out


@st.cache_data(ttl=60)
def load_variance(path: str = str(XLSX_FILE)) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb["Variance"]
    out = {}
    for key, row in VARIANCE_ROWS.items():
        out[key] = {
            "spring": ws.cell(row=row, column=3).value or 0,
            "fall":   ws.cell(row=row, column=4).value or 0,
            "var_eur":ws.cell(row=row, column=5).value or 0,
            "var_pct":ws.cell(row=row, column=6).value or 0,
        }
    wb.close()
    return out
